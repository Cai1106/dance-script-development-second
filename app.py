import os
import re
import shutil
import difflib
import threading
import tkinter as tk
from tkinter import filedialog
import openpyxl
import uvicorn
import asyncio
from fastapi.responses import HTMLResponse
from fastapi import FastAPI, WebSocket, Request
from fastapi.staticfiles import StaticFiles
from fastapi.templating import Jinja2Templates
from fastapi.middleware.cors import CORSMiddleware
import webview
from webview.dom import DOMEventHandler
import time
import socket
import urllib.request
from urllib.error import URLError


# 尝试导入简繁转换库
try:
    import zhconv
    HAS_ZHCONV = True
except ImportError:
    HAS_ZHCONV = False
    print("未安装 zhconv 库，简繁体转换功能可能受限。")

# ==========================================
# 默认搜索磁盘配置区域（小白请看这里修改）
# ==========================================
# 这里是设置默认去哪里找音乐的路径列表。
# 你可以增加或删除路径。注意路径要用引号括起来，如果有反斜杠请用双斜杠（\\）或者在前面加r。
# 比如：r"D:\舞蹈音乐" 或者 "E:\\所有音乐"
DEFAULT_SEARCH_PATHS = [
    r"E:/",
    
    # 在这里继续添加你的默认路径，记得用逗号隔开...
]
# ==========================================

def normalize_string(s):
    """
    处理字符串：
    1. 转小写（忽略英文大小写）
    2. 去除所有空格
    3. 如果有简繁体转换库，统一转为简体中文以便比较
    """
    if not isinstance(s, str):
        return ""
    s = s.lower().replace(" ", "").replace("\u3000", "")
    if HAS_ZHCONV:
        s = zhconv.convert(s, 'zh-hans')
    return s

def extract_dance_name(raw_name):
    """
    提取舞名：
    1. 去掉开头的数字和点，如 "1.阿路娜" -> "阿路娜"
    2. 去掉括号及其里面的内容，如 "阿路娜(罗)" -> "阿路娜"
    """
    if not isinstance(raw_name, str):
        return ""
    # 去除前面的序号（例如 "1.", "12、", " 3." 等）
    name = re.sub(r'^\s*\d+[\.、\s]+', '', raw_name)
    # 去除括号及其内部内容，支持中文括号和英文括号
    name = re.sub(r'[（\(].*?[）\)]', '', name)
    # 去除两端空格
    return name.strip()

app = FastAPI()

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# Ensure directories exist
os.makedirs("static", exist_ok=True)
os.makedirs("templates", exist_ok=True)

# Mount static files
app.mount("/static", StaticFiles(directory="static"), name="static")
app.mount("/images", StaticFiles(directory="images"), name="images")

templates = Jinja2Templates(directory="templates")


@app.get("/")
async def read_root(request: Request):
    template = templates.get_template("index.html")
    content = template.render(request=request)
    return HTMLResponse(content=content)
     

@app.get("/api/init")
async def init_data():
    desktop_path = os.path.join(os.path.expanduser("~"), "Desktop")
    valid_paths = [p for p in DEFAULT_SEARCH_PATHS if os.path.exists(p)]
    return {
        "default_paths": valid_paths,
        "default_save_path": desktop_path
    }

def open_file_dialog(dialog_func, **kwargs):
    root = tk.Tk()
    root.withdraw()
    root.attributes('-topmost', True)
    result = dialog_func(**kwargs)
    root.destroy()
    return result

@app.get("/api/select_folder")
async def select_folder():
    folder = await asyncio.to_thread(open_file_dialog, filedialog.askdirectory, title="选择文件夹")
    return {"path": folder}

@app.get("/api/select_excel")
async def select_excel():
    file = await asyncio.to_thread(open_file_dialog, filedialog.askopenfilename, title="选择Excel文件", filetypes=[("Excel files", "*.xlsx *.xls")])
    return {"path": file}

@app.get("/api/open_doc")
async def open_doc():
    doc_path = os.path.join(os.path.dirname(__file__), "相关文档", "说明文档.docx")
    # doc_path = os.path.abspath("说明文档.docx")
    if os.path.exists(doc_path):
        try:
            os.startfile(doc_path)
            return {"status": "success"}
        except Exception as e:
            return {"status": "error", "msg": str(e)}
    else:
        return {"status": "error", "msg": "未找到 说明文档.docx"}

@app.post("/api/delete_excel")
async def delete_excel(request: Request):
    data = await request.json()
    path = data.get("path")
    if not path or not os.path.exists(path):
        return {"status": "error", "msg": "文件不存在或路径为空"}
    
    import subprocess
    # 处理路径中的单引号
    safe_path = path.replace("'", "''")
    ps_command = f"Add-Type -AssemblyName Microsoft.VisualBasic; [Microsoft.VisualBasic.FileIO.FileSystem]::DeleteFile('{safe_path}', 'OnlyErrorDialogs', 'SendToRecycleBin')"
    try:
        subprocess.run(["powershell", "-NoProfile", "-Command", ps_command], check=True, creationflags=subprocess.CREATE_NO_WINDOW)
        return {"status": "success"}
    except Exception as e:
        return {"status": "error", "msg": str(e)}

@app.post("/api/clean_excel")
async def clean_excel(request: Request):
    data = await request.json()
    path = data.get("path")
    if not path or not os.path.exists(path):
        return {"status": "error", "msg": "Excel文件不存在或路径为空"}
    
    try:
        import tempfile
        
        # 为了防止文件被占用，先复制一份
        temp_excel_path = os.path.join(tempfile.gettempdir(), "temp_clean_excel.xlsx")
        shutil.copy2(path, temp_excel_path)
        
        wb = openpyxl.load_workbook(temp_excel_path)
        cleaned_count = 0
        
        # 遍历所有的sheet和所有的单元格
        for sheet in wb.worksheets:
            for row in sheet.iter_rows():
                for cell in row:
                    if isinstance(cell.value, str):
                        original = cell.value
                        # 去除括号及其内部内容，支持中文括号和英文括号
                        cleaned = re.sub(r'[（\(].*?[）\)]', '', original)
                        if cleaned != original:
                            cell.value = cleaned
                            cleaned_count += 1
                            
        # 保存并替换原文件
        wb.save(temp_excel_path)
        wb.close()
        
        shutil.copy2(temp_excel_path, path)
        os.remove(temp_excel_path)
        
        return {"status": "success", "msg": f"成功清理了 {cleaned_count} 个带有国家名称的舞名！"}
    except PermissionError:
        return {"status": "error", "msg": "清理失败：文件被拒绝访问。请确保您已经关闭了该Excel文件！"}
    except Exception as e:
        return {"status": "error", "msg": f"清理失败: {str(e)}"}

@app.websocket("/ws")
async def websocket_endpoint(websocket: WebSocket):
    await websocket.accept()
    
    music_files_cache = {}
    
    try:
        data = await websocket.receive_json()
        action = data.get("action")
        
        if action == "start":
            mode = data.get("mode", "party")
            export_type = data.get("export_type", "copy") # 获取导出方式，默认快捷方式
            paths = data.get("paths", [])
            excel_path = data.get("excel_path", "")
            save_path = data.get("save_path", "")
            
            if not excel_path or not os.path.exists(excel_path):
                await websocket.send_json({"type": "error", "msg": "Excel文件无效"})
                return
            if not paths:
                await websocket.send_json({"type": "error", "msg": "未指定音乐搜索路径"})
                return
            
            await websocket.send_json({"type": "progress", "msg": "正在扫描磁盘中的音乐文件，请稍候...", "val": 10})
            
            for path in paths:
                if not os.path.exists(path):
                    continue
                for root, dirs, files in os.walk(path):
                    for file in files:
                        if file.lower().endswith(('.mp3', '.wma')):
                            name_without_ext = os.path.splitext(file)[0]
                            norm_name = normalize_string(name_without_ext)
                            full_path = os.path.join(root, file)
                            if norm_name not in music_files_cache:
                                music_files_cache[norm_name] = full_path
                                
            await websocket.send_json({"type": "progress", "msg": f"扫描完成，共找到 {len(music_files_cache)} 首音乐。", "val": 30})
            
            await websocket.send_json({"type": "progress", "msg": "正在读取Excel表格...", "val": 40})
            try:
                # 尝试复制一份临时的 Excel 文件来读取，防止原文件被 Excel 软件锁住导致拒绝访问
                import tempfile
                temp_excel_path = os.path.join(tempfile.gettempdir(), "temp_dance_excel.xlsx")
                try:
                    shutil.copy2(excel_path, temp_excel_path)
                    excel_to_read = temp_excel_path
                except Exception as copy_e:
                    print(f"复制临时文件失败，尝试直接读取: {copy_e}")
                    excel_to_read = excel_path

                wb = openpyxl.load_workbook(excel_to_read, data_only=True)
                sheet = wb.active
                folder_name = str(sheet['A4'].value or "未知舞单").strip()
                wb.close()
                
                # 读取完毕后尝试删除临时文件
                if excel_to_read == temp_excel_path:
                    try:
                        os.remove(temp_excel_path)
                    except:
                        pass
                        
            except PermissionError:
                await websocket.send_json({"type": "error", "msg": "读取Excel失败：文件被拒绝访问。请确保您已经关闭了该Excel文件！"})
                return
            except Exception as e:
                await websocket.send_json({"type": "error", "msg": f"读取Excel失败: {str(e)}"})
                return
                
            dance_names = []
            
            def add_cells(col_letter, start_row, end_row, suffix="", start_index=1):
                idx = start_index
                for row in range(start_row, end_row + 1):
                    val = sheet[f"{col_letter}{row}"].value
                    if val:
                        clean_name = extract_dance_name(str(val))
                        if clean_name:
                            dance_names.append((clean_name, idx, suffix))
                            idx += 1
                return idx
                            
            if mode == "party":
                idx = add_cells('B', 8, 32, start_index=1)
                idx = add_cells('C', 8, 32, start_index=idx)
                add_cells('D', 8, 27, start_index=idx)
            else:
                # 日常社活舞单规则
                # 1. 复习舞单 B8到B17 (10首)，后缀“复习”
                add_cells('B', 8, 17, suffix="复习", start_index=1)
                # 2. 教学舞单 C8到C10 (3首)，后缀“教学”
                add_cells('C', 8, 10, suffix="教学", start_index=1)
                # 3. 正常舞单 B21到B30, C21到C30, D21到D30 (30首)，连续编号无后缀
                idx = add_cells('B', 21, 30, start_index=1)
                idx = add_cells('C', 21, 30, start_index=idx)
                add_cells('D', 21, 30, start_index=idx)
                # 4. 备用舞单 D8到D17 (10首)，后缀“备用”
                add_cells('D', 8, 17, suffix="备用", start_index=1)
                
            if not dance_names:
                await websocket.send_json({"type": "error", "msg": "未在表格指定位置找到任何舞名。"})
                return
                
            save_dir = os.path.join(save_path, folder_name)
            if not os.path.exists(save_dir):
                os.makedirs(save_dir)
                
            not_found = []
            m3u_content = [] # 用于存储歌单的绝对路径
            found_music_list = [] # 用于存储找到的音乐信息 (target_name, index, suffix, found_path)
            total_dances = len(dance_names)
            
            for i, (target_name, index, suffix) in enumerate(dance_names):
                progress_val = 40 + int(50 * (i / total_dances))
                await websocket.send_json({"type": "progress", "msg": f"正在处理: {target_name}", "val": progress_val})
                
                norm_target = normalize_string(target_name)
                found_path = None
                
                for cache_name, file_path in music_files_cache.items():
                    if norm_target == cache_name:
                        found_path = file_path
                        break
                        
                if not found_path:
                    for cache_name, file_path in music_files_cache.items():
                        if norm_target in cache_name:
                            found_path = file_path
                            break
                            
                if not found_path:
                    all_names = list(music_files_cache.keys())
                    close_matches = difflib.get_close_matches(norm_target, all_names, n=3, cutoff=0.4)
                    
                    if close_matches:
                        for match in close_matches:
                            real_file_path = music_files_cache[match]
                            filename_only = os.path.basename(real_file_path)
                            
                            # Ask user via websocket
                            await websocket.send_json({
                                "type": "ask",
                                "target": target_name,
                                "match": filename_only
                            })
                            
                            # Wait for answer
                            resp = await websocket.receive_json()
                            if resp.get("answer") == "yes":
                                found_path = real_file_path
                                break
                            elif resp.get("answer") == "cancel":
                                break
                                
                if found_path:
                    # 将找到的原始文件绝对路径添加到歌单列表中
                    m3u_content.append(found_path)
                    found_music_list.append((target_name, index, suffix, found_path))
                else:
                    not_found.append(target_name)
                    
            # ---------------------------------------------------------
            # 根据用户选择的导出方式执行不同逻辑
            # ---------------------------------------------------------
            if export_type == "m3u":
                # 方案 1: 仅生成 m3u
                m3u_file_path = os.path.join(save_dir, f"{folder_name}.m3u")
                try:
                    with open(m3u_file_path, 'w', encoding='utf-8') as f:
                        f.write("#EXTM3U\n")
                        for p in m3u_content:
                            f.write(f"{p}\n")
                    try:
                        os.startfile(m3u_file_path)
                    except Exception:
                        pass
                    await websocket.send_json({
                        "type": "done",
                        "msg": f"M3U歌单生成完成！保存在：{save_dir}\n已尝试自动打开，您也可以手动拖入播放器。",
                        "not_found": not_found
                    })
                except Exception as e:
                    await websocket.send_json({"type": "error", "msg": f"生成 M3U 歌单失败: {e}"})
                    
            elif export_type == "copy":
                # 方案 2: 真实复制文件
                copy_dir = os.path.join(save_dir, f"{folder_name}_实体音乐文件")
                if not os.path.exists(copy_dir):
                    os.makedirs(copy_dir)
                
                copy_count = 0
                for target_name, index, suffix, found_path in found_music_list:
                    ext = os.path.splitext(found_path)[1]
                    if suffix:
                        new_file_name = f"{index:02d} - {target_name}（{suffix}）{ext}"
                    else:
                        new_file_name = f"{index:02d} - {target_name}{ext}"
                    dest_path = os.path.join(copy_dir, new_file_name)
                    try:
                        shutil.copy2(found_path, dest_path)
                        copy_count += 1
                    except Exception as e:
                        print(f"复制失败 {found_path}: {e}")
                
                try:
                    os.startfile(copy_dir)
                except Exception:
                    pass
                    
                await websocket.send_json({
                    "type": "done",
                    "msg": f"成功复制了 {copy_count} 首音乐！\n保存在：{copy_dir}",
                    "not_found": not_found
                })
                
            else:
                # 方案 3 (默认): 生成快捷方式
                try:
                    # 尝试使用 win32com (需要 pywin32 库)
                    try:
                        import win32com.client
                        import pythoncom
                        use_win32com = True
                    except ImportError:
                        print("未找到 win32com，改用 VBScript 方案...")
                        use_win32com = False
                    
                    print("开始生成快捷方式...")
                    
                    # 为每首歌创建一个快捷方式，保持文件名里的排序
                    shortcuts_dir = os.path.join(save_dir, f"{folder_name}_快捷方式_可拖入酷狗")
                    print(f"准备创建快捷方式文件夹: {shortcuts_dir}")
                    
                    if not os.path.exists(shortcuts_dir):
                        os.makedirs(shortcuts_dir)
                        
                    if use_win32com:
                        # 必须在线程中初始化 COM 库
                        pythoncom.CoInitialize()
                        shell = win32com.client.Dispatch("WScript.Shell")
                        
                    shortcut_count = 0
                    for target_name, index, suffix, found_path in found_music_list:
                        if suffix:
                            lnk_name = f"{index:02d} - {target_name}（{suffix}）.lnk"
                        else:
                            lnk_name = f"{index:02d} - {target_name}.lnk"
                            
                        lnk_path = os.path.join(shortcuts_dir, lnk_name)
                        
                        try:
                            if use_win32com:
                                shortcut = shell.CreateShortCut(lnk_path)
                                shortcut.Targetpath = found_path
                                shortcut.WindowStyle = 1 # Normal
                                shortcut.save()
                            else:
                                # VBScript 方案作为完美备胎，不需要任何第三方库
                                vbs_path = os.path.join(shortcuts_dir, "temp_make_shortcut.vbs")
                                # VBScript 字符串中如果路径有双引号需要处理，但 Windows 路径通常不含双引号
                                vbs_code = f'''
                                Set Shell = CreateObject("WScript.Shell")
                                Set Link = Shell.CreateShortcut("{lnk_path}")
                                Link.TargetPath = "{found_path}"
                                Link.Save
                                '''
                                # 使用 utf-16 编码并带 BOM，VBScript/cscript 对此支持较好，能处理更多特殊字符
                                with open(vbs_path, "w", encoding="utf-16") as vbs_file:
                                    vbs_file.write(vbs_code)
                                os.system(f'cscript //nologo "{vbs_path}"')
                                if os.path.exists(vbs_path):
                                    os.remove(vbs_path)
                                    
                            shortcut_count += 1
                        except Exception as e_lnk:
                            print(f"创建单个快捷方式失败 [{lnk_name}]: {e_lnk}")
                            
                    if use_win32com:
                        # 释放 COM 库
                        pythoncom.CoUninitialize()
                        
                    print(f"成功生成了 {shortcut_count} 个快捷方式")
                    
                    # 生成快捷方式后，打开这个文件夹
                    try:
                        os.startfile(shortcuts_dir)
                    except Exception as e:
                        print(f"自动打开文件夹失败: {e}")
                        pass
                        
                    await websocket.send_json({
                        "type": "done",
                        "msg": f"快捷方式生成完成！保存在：{save_dir}\n已自动打开文件夹，请全选拖入酷狗音乐。",
                        "not_found": not_found
                    })
                        
                except Exception as e:
                    print(f"生成快捷方式整体流程失败: {e}")
                    # 将错误信息发送到前端，这样你就能直接在界面上看到了
                    await websocket.send_json({"type": "error", "msg": f"生成快捷方式失败，但歌单已生成。错误信息: {e}"})
            
    except Exception as e:
        await websocket.send_json({"type": "error", "msg": str(e)})

def start_server(port):
    uvicorn.run("app:app", host="127.0.0.1", port=port, log_level="error")

def get_free_port():
    s = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
    s.bind(('127.0.0.1', 0))
    port = s.getsockname()[1]
    s.close()
    return port

def wait_for_server(port):
    """轮询检查服务器是否启动完毕"""
    url = f"http://127.0.0.1:{port}/api/init"
    for _ in range(30): # 最多等待3秒
        try:
            urllib.request.urlopen(url)
            return True
        except URLError:
            time.sleep(0.1)
    return False

import json

def bind_drag_drop(window):
    def on_loaded():
        def on_drop(e):
            files = e.get('dataTransfer', {}).get('files', [])
            if files:
                file_path = files[0].get('pywebviewFullPath')
                if file_path:
                    if file_path.lower().endswith('.xlsx') or file_path.lower().endswith('.xls'):
                        # 使用 json.dumps 确保路径被正确转义为 JS 字符串
                        js_code = f"if(window.updateExcelPath) window.updateExcelPath({json.dumps(file_path)})"
                        window.evaluate_js(js_code)
                    else:
                        window.evaluate_js("alert('请拖入 Excel 文件 (.xlsx 或 .xls)')")

        try:
            window.dom.document.events.drop += DOMEventHandler(on_drop, False, False)
            print("拖拽事件绑定成功")
        except Exception as e:
            print(f"绑定拖拽事件失败: {e}")

    window.events.loaded += on_loaded

if __name__ == "__main__":
    port = get_free_port()
    
    # 启动 FastAPI 后端服务
    server_thread = threading.Thread(target=start_server, args=(port,), daemon=True)
    server_thread.start()
    
    # 智能等待服务启动，而不是硬编码死等
    wait_for_server(port)
    
    # 启动本地客户端窗口
    window = webview.create_window(
        title="舞蹈音乐提取器", 
        url=f"http://127.0.0.1:{port}", 
        width=850, 
        height=780,
        resizable=False,  # 禁止调整大小以防止出现滚动条
        min_size=(850, 780),
        background_color='#f3f4f6' # 窗口背景色匹配页面，防止白屏闪烁
    )
    webview.start(bind_drag_drop, window)
    
    # 界面关闭后强制结束整个进程（包括后端的 Uvicorn）
    os._exit(0)

